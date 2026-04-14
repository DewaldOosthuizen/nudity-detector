"""
Report management for the Nudity Detector application.
Handles report file I/O, Excel generation, and session persistence.
Single responsibility: Report and session file management only.
"""

import base64
import json
import logging
import os
from io import BytesIO
from typing import List, Dict, Any, Optional

import openpyxl

try:
    from PIL import Image
except ImportError:
    Image = None

try:
    from openpyxl.drawing.image import Image as XLImage
except ImportError:
    XLImage = None

from ..core import constants
from ..core.models import ReportEntry, SessionState, ScanConfig


class ReportManager:
    """Manages report file operations (reading/writing Excel, sessions)."""

    @staticmethod
    def get_report_path(report_dir: str = constants.DEFAULT_REPORT_DIR) -> str:
        """Get standard report file path.

        Args:
            report_dir: Report directory path

        Returns:
            Full path to report file
        """
        return os.path.join(report_dir, constants.REPORT_FILE_NAME)

    @staticmethod
    def get_session_path(report_file_path: str) -> str:
        """Get session file path corresponding to report file.

        Args:
            report_file_path: Path to report Excel file

        Returns:
            Path to corresponding session JSON file
        """
        base_name, _ = os.path.splitext(report_file_path)
        return f'{base_name}{constants.SESSION_FILE_SUFFIX}'

    @staticmethod
    def validate_report_dir(report_dir: str) -> tuple[bool, str]:
        """Validate report directory is writable and safe.

        Args:
            report_dir: Path to report directory

        Returns:
            Tuple of (valid, error_message)
        """
        if not report_dir:
            return False, 'Report directory cannot be empty'

        if report_dir in constants.SYSTEM_PROTECTED_DIRS:
            return False, f'Cannot write reports to system directory: {report_dir}'

        try:
            os.makedirs(report_dir, exist_ok=True)
            test_file = os.path.join(report_dir, '.test')
            with open(test_file, 'w') as f:
                f.write('test')
            os.remove(test_file)
            return True, ''
        except (OSError, IOError) as e:
            return False, f'Report directory not writable: {e}'

    @staticmethod
    def load_entries(file_path: str) -> List[ReportEntry]:
        """Load report entries from Excel file.

        Supports both old schema (Copied File) and new schema (Thumbnail).

        Args:
            file_path: Path to Excel report file

        Returns:
            List of ReportEntry objects
        """
        if not os.path.exists(file_path):
            return []

        try:
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active

            if sheet.max_row < 2:  # No data rows
                return []

            normalized_headers = {
                str(cell.value).strip(): idx
                for idx, cell in enumerate(sheet[1])
                if cell.value
            }

            entries = []
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if not row or not row[0]:  # Skip empty rows
                    continue

                try:
                    # Map columns by header name
                    entry_data = {
                        constants.RESULT_FIELD_FILE: row[normalized_headers.get('File', 0)] or '',
                        constants.RESULT_FIELD_MEDIA_TYPE: row[normalized_headers.get('Media Type', 1)] or constants.MEDIA_TYPE_UNKNOWN,
                        constants.RESULT_FIELD_MODEL: row[normalized_headers.get('Model', 2)] or '',
                        constants.RESULT_FIELD_THRESHOLD: row[normalized_headers.get('Threshold Percent', 3)] or constants.DEFAULT_THRESHOLD_PERCENT,
                        constants.RESULT_FIELD_CONFIDENCE: row[normalized_headers.get('Confidence Percent', 4)] or 0.0,
                        constants.RESULT_FIELD_NUDITY: bool(row[normalized_headers.get('Nudity Detected', 5)]),
                        constants.RESULT_FIELD_CLASSES: row[normalized_headers.get('Detected Classes', 6)] or '[]',
                        constants.RESULT_FIELD_THUMBNAIL: row[normalized_headers.get('Thumbnail', 7)] or '',
                        constants.RESULT_FIELD_DATE: row[normalized_headers.get('Date Classified', 8)] or '',
                    }
                    entries.append(ReportEntry.from_dict(entry_data))
                except (IndexError, ValueError, KeyError) as e:
                    logging.warning('Skipping malformed report row %d: %s', row_idx, e)
                    continue

            return entries
        except Exception as e:
            logging.error('Failed to load report entries from %s: %s', file_path, e)
            return []

    @staticmethod
    def save_entries(entries: List[ReportEntry], file_path: str) -> bool:
        """Save report entries to Excel file.

        Args:
            entries: List of ReportEntry objects
            file_path: Path to save Excel file

        Returns:
            True if successful
        """
        try:
            os.makedirs(os.path.dirname(file_path) or '.', exist_ok=True)

            if os.path.exists(file_path):
                workbook = openpyxl.load_workbook(file_path)
                sheet = workbook.active
            else:
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                sheet.title = 'Nudity Report'

            # Clear old data and write headers
            sheet.delete_rows(1, sheet.max_row or 1)
            sheet.append(constants.REPORT_HEADERS)

            # Write entries
            for entry in entries:
                sheet.append(entry.to_row())

            # Embed thumbnails
            ReportManager._embed_thumbnails(sheet, entries)

            workbook.save(file_path)
            logging.info('Report saved to %s', file_path)
            return True
        except Exception as e:
            logging.error('Failed to save report to %s: %s', file_path, e)
            return False

    @staticmethod
    def _embed_thumbnails(sheet, entries: List[ReportEntry]) -> None:
        """Embed thumbnail images into report sheet.

        Args:
            sheet: openpyxl worksheet
            entries: List of ReportEntry objects
        """
        if XLImage is None or Image is None:
            logging.debug('Cannot embed thumbnails: openpyxl Image or PIL not available')
            return

        try:
            # Find Thumbnail column
            headers = [cell.value for cell in sheet[1]]
            if 'Thumbnail' not in headers:
                return

            thumbnail_col_idx = headers.index('Thumbnail') + 1
            thumbnail_col = openpyxl.utils.get_column_letter(thumbnail_col_idx)
            sheet.column_dimensions[thumbnail_col].width = 15

            # Embed thumbnails
            for row_idx, entry in enumerate(entries, start=2):
                if not entry.thumbnail:
                    continue

                try:
                    thumbnail_data = base64.b64decode(entry.thumbnail)
                    img = Image.open(BytesIO(thumbnail_data))

                    img_bytes = BytesIO()
                    img.save(img_bytes, format='PNG')
                    img_bytes.seek(0)

                    cell_ref = f'{thumbnail_col}{row_idx}'
                    xl_image = XLImage(img_bytes)
                    xl_image.width = 100
                    xl_image.height = 100
                    sheet.add_image(xl_image, cell_ref)

                    sheet.row_dimensions[row_idx].height = 105
                except Exception as e:
                    logging.debug('Failed to embed thumbnail in %s%d: %s', thumbnail_col, row_idx, e)
                    continue
        except Exception as e:
            logging.warning('Failed to embed thumbnails: %s', e)

    @staticmethod
    def save_session(session_state: SessionState, report_file_path: str) -> bool:
        """Save session state to JSON file.

        Args:
            session_state: SessionState object
            report_file_path: Path to corresponding report file

        Returns:
            True if successful
        """
        try:
            session_path = ReportManager.get_session_path(report_file_path)
            os.makedirs(os.path.dirname(session_path) or '.', exist_ok=True)

            with open(session_path, 'w', encoding='utf-8') as f:
                json.dump(session_state.to_dict(), f, ensure_ascii=False, indent=2)

            logging.info('Session saved to %s', session_path)
            return True
        except Exception as e:
            logging.error('Failed to save session: %s', e)
            return False

    @staticmethod
    def load_session(file_path: str) -> SessionState:
        """Load session state from file (JSON or Excel).

        Args:
            file_path: Path to session JSON or report Excel

        Returns:
            SessionState object
        """
        # Try JSON file first
        session_path = file_path if file_path.endswith('.json') else ReportManager.get_session_path(file_path)

        if os.path.exists(session_path):
            try:
                with open(session_path, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    return SessionState.from_dict(data)
            except (OSError, json.JSONDecodeError) as e:
                logging.warning('Failed to load session from %s: %s', session_path, e)

        # Try embedded session in Excel
        if os.path.exists(file_path) and file_path.endswith(constants.XLSX_EXTENSION):
            try:
                workbook = openpyxl.load_workbook(file_path)
                if 'Session' in workbook.sheetnames:
                    raw_session = workbook['Session']['A1'].value
                    if raw_session:
                        try:
                            data = json.loads(raw_session)
                            return SessionState.from_dict(data)
                        except json.JSONDecodeError as e:
                            logging.warning('Could not parse embedded session: %s', e)

                # Fallback: create from report entries
                entries = ReportManager.load_entries(file_path)
                return SessionState(results=[e for e in entries if e.nudity_detected])
            except Exception as e:
                logging.warning('Failed to load session from %s: %s', file_path, e)

        # Default empty session
        return SessionState()

    @staticmethod
    def create_demo_session(report_path: str, session_state: SessionState) -> bool:
        """Create or update session data in Excel workbook.

        Args:
            report_path: Path to Excel report
            session_state: SessionState to embed

        Returns:
            True if successful
        """
        try:
            workbook = openpyxl.load_workbook(report_path) if os.path.exists(report_path) else openpyxl.Workbook()

            if 'Session' not in workbook.sheetnames:
                workbook.create_sheet('Session')

            session_sheet = workbook['Session']
            session_sheet.delete_rows(1, session_sheet.max_row or 1)
            session_sheet['A1'] = json.dumps(session_state.to_dict(), ensure_ascii=False, indent=2)

            workbook.save(report_path)
            return True
        except Exception as e:
            logging.error('Failed to create demo session: %s', e)
            return False
