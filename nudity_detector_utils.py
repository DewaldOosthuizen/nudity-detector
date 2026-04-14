"""Nudity detection utility functions.

This module provides utilities for:
- Image and video file classification using AI models (NudeNet, DeepStack)
- Thumbnail generation (base64-encoded) for detected items
- Report generation and persistence (Excel, JSON)
- Thread-safe batch processing with worker pools
- Timeout protection for long-running detections

Thread Safety:
- nudity_report list is protected by report_lock
- Worker threads are non-daemon to ensure proper cleanup
- All thread operations have explicit join() with timeout

Schema:
- Report entries include: file, media_type, model_name, confidence, nudity_detected, 
  detected_classes, thumbnail, date_classified
- Thumbnails are stored as base64-encoded PNG strings for portability
"""

import base64
import json
import logging
import os
import shutil
import subprocess
import sys
from datetime import datetime
from io import BytesIO
from queue import Queue
from threading import Lock, Thread
from typing import Optional, Tuple

import openpyxl

try:
    from PIL import Image
except ImportError:
    Image = None

try:
    import cv2
except ImportError:
    cv2 = None

try:
    from send2trash import send2trash
except ImportError:
    send2trash = None




def validate_report_dir(report_dir: str) -> Tuple[bool, str]:
    """Validate report directory is writable and safe.
    
    Args:
        report_dir: Path to report directory
        
    Returns:
        Tuple of (valid: bool, error_message: str)
    """
    if not report_dir:
        return False, 'Report directory cannot be empty'
    
    # Prevent writing to system directories
    if report_dir in ['/', '/etc', '/sys', '/dev', '/proc', '/root']:
        return False, f'Cannot write reports to system directory: {report_dir}'
    
    try:
        os.makedirs(report_dir, exist_ok=True)
        # Try to write a test file
        test_file = os.path.join(report_dir, '.test')
        with open(test_file, 'w') as f:
            f.write('test')
        os.remove(test_file)
        return True, ''
    except (OSError, IOError) as e:
        return False, f'Report directory not writable: {e}'

image_extensions = {'.png', '.jpg', '.jpeg', '.gif', '.bmp', '.webp', '.tiff'}
video_extensions = {'.mp4', '.avi', '.mkv', '.mov', '.vob', '.wmv', '.flv', '.3gp', '.webm'}
REPORT_HEADERS = [
    'File',
    'Media Type',
    'Model',
    'Threshold Percent',
    'Confidence Percent',
    'Nudity Detected',
    'Detected Classes',
    'Thumbnail',
    'Date Classified',
]
SESSION_VERSION = 1
DEFAULT_REPORT_DIR = 'reports'

nudity_report = []
report_lock = Lock()


def get_report_path(report_dir=DEFAULT_REPORT_DIR):
    return os.path.join(report_dir, 'nudity_report.xlsx')


def get_session_path(report_file_path):
    base_name, _ = os.path.splitext(report_file_path)
    return f'{base_name}_session.json'


def normalize_threshold(threshold_value):
    if threshold_value is None:
        return 0.6

    threshold = float(threshold_value)
    if threshold > 1:
        threshold /= 100.0
    return max(0.0, min(threshold, 1.0))


def threshold_to_percent(threshold_value):
    return round(normalize_threshold(threshold_value) * 100, 2)


def detect_media_type(file_path):
    _, ext = os.path.splitext(file_path)
    ext = ext.lower()
    if ext in image_extensions:
        return 'image'
    if ext in video_extensions:
        return 'video'
    return 'unknown'



def generate_image_thumbnail(file_path: str, size: Tuple[int, int] = (100, 100)) -> Optional[str]:
    """Generate base64-encoded thumbnail from image file.
    
    Args:
        file_path: Path to image file
        size: Thumbnail size as (width, height) tuple
        
    Returns:
        Base64-encoded PNG thumbnail string, or None if generation fails
    """
    if Image is None:
        logging.debug('PIL not available for thumbnail generation: %s', file_path)
        return None
        
    try:
        with Image.open(file_path) as img:
            img.thumbnail(size, Image.Resampling.LANCZOS)
            # Convert to RGB if needed (e.g., RGBA or grayscale)
            if img.mode != 'RGB':
                img = img.convert('RGB')
            
            # Encode to base64
            buffer = BytesIO()
            img.save(buffer, format='PNG')
            buffer.seek(0)
            encoded = base64.b64encode(buffer.getvalue()).decode('utf-8')
            return encoded
    except Exception as e:
        logging.warning('Failed to generate image thumbnail for %s: %s', file_path, e)
        return None


def generate_video_thumbnail(file_path: str, size: Tuple[int, int] = (100, 100)) -> Optional[str]:
    """Generate base64-encoded thumbnail from video file at 25% progress.
    
    Args:
        file_path: Path to video file
        size: Thumbnail size as (width, height) tuple
        
    Returns:
        Base64-encoded PNG thumbnail string, or None if generation fails
    """
    if cv2 is None:
        logging.debug('OpenCV not available for video thumbnail generation: %s', file_path)
        return None
        
    try:
        cap = cv2.VideoCapture(file_path)
        if not cap.isOpened():
            logging.warning('Could not open video file: %s', file_path)
            return None
            
        total_frames = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
        if total_frames <= 0:
            logging.warning('Video has no frames: %s', file_path)
            cap.release()
            return None
        
        # Extract frame at 25% progress
        frame_index = max(0, int(total_frames * 0.25))
        cap.set(cv2.CAP_PROP_POS_FRAMES, frame_index)
        
        ret, frame = cap.read()
        cap.release()
        
        if not ret or frame is None:
            logging.warning('Could not extract frame from video: %s', file_path)
            return None
        
        # Convert BGR to RGB and create thumbnail
        frame_rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        img = Image.fromarray(frame_rgb)
        img.thumbnail(size, Image.Resampling.LANCZOS)
        
        # Encode to base64
        buffer = BytesIO()
        img.save(buffer, format='PNG')
        buffer.seek(0)
        encoded = base64.b64encode(buffer.getvalue()).decode('utf-8')
        return encoded
    except Exception as e:
        logging.warning('Failed to generate video thumbnail for %s: %s', file_path, e)
        return None


def get_thumbnail(file_path: str, media_type: Optional[str] = None) -> Optional[str]:
    """Get thumbnail for image or video file.
    
    Args:
        file_path: Path to media file
        media_type: 'image' or 'video', auto-detected if None
        
    Returns:
        Base64-encoded thumbnail or None if unavailable
    """
    if not os.path.exists(file_path):
        return None
        
    media_type = media_type or detect_media_type(file_path)
    
    if media_type == 'image':
        return generate_image_thumbnail(file_path)
    elif media_type == 'video':
        return generate_video_thumbnail(file_path)
    
    return None


def open_file(file_path: str) -> Tuple[bool, str]:
    """Open file directly (not parent directory).
    
    Args:
        file_path: Path to file to open
        
    Returns:
        Tuple of (success: bool, error_message: str)
    """
    if not os.path.exists(file_path):
        error_msg = f'File does not exist: {file_path}'
        logging.warning(error_msg)
        return False, error_msg
    
    try:
        if sys.platform == 'win32':
            os.startfile(file_path)
        elif sys.platform == 'darwin':
            subprocess.run(['open', file_path], check=False)
        else:
            subprocess.run(['xdg-open', file_path], check=False)
        return True, ''
    except Exception as error:
        error_msg = f'Could not open file {file_path}: {error}'
        logging.error(error_msg)
        return False, error_msg


def detect_with_timeout(detector, file_path: str, timeout_seconds: int = 60) -> Optional[list]:
    """Wrap NudeNet detection with timeout.
    
    Args:
        detector: NudeNet detector instance  
        file_path: Path to media file
        timeout_seconds: Timeout in seconds
        
    Returns:
        Detection results or None if timeout
        
    Raises:
        TimeoutError: If detection exceeds timeout
    """
    from threading import Timer
    
    result_container = [None]
    exception_container = [None]
    
    def run_detection():
        try:
            result_container[0] = detector.detect(file_path)
        except Exception as e:
            exception_container[0] = e
    
    thread = Thread(target=run_detection, daemon=False)
    thread.start()
    thread.join(timeout=timeout_seconds)
    
    if thread.is_alive():
        logging.error('Detection timeout for file: %s after %d seconds', file_path, timeout_seconds)
        raise TimeoutError(f'Detection timeout for {file_path}')
    
    if exception_container[0]:
        raise exception_container[0]
    
    return result_container[0]

def make_scan_config(source_folder='', model_name='nudenet', threshold_percent=60, theme_mode='system'):
    return {
        'theme_mode': theme_mode,
        'source_folder': source_folder,
        'model_name': model_name,
        'threshold_percent': threshold_to_percent(threshold_percent),
    }


def create_session_state(scan_config=None, results=None):
    return {
        'version': SESSION_VERSION,
        'saved_at': datetime.now().isoformat(timespec='seconds'),
        'scan_config': scan_config or make_scan_config(),
        'results': results or [],
    }


def reset_nudity_report():
    with report_lock:
        nudity_report.clear()


def replace_nudity_report(entries):
    with report_lock:
        nudity_report.clear()
        nudity_report.extend(entries)


def get_detected_results(report_data=None):
    data = report_data if report_data is not None else nudity_report
    return [entry for entry in data if entry.get('nudity_detected')]


def parse_detected_classes(raw_result):
    if raw_result is None:
        return []

    parsed_result = raw_result
    if isinstance(raw_result, str):
        try:
            parsed_result = json.loads(raw_result)
        except json.JSONDecodeError:
            parsed_result = raw_result

    if isinstance(parsed_result, list):
        labels = []
        for item in parsed_result:
            if isinstance(item, dict):
                label = item.get('class') or item.get('label') or item.get('frame') or item.get('name')
                if label:
                    labels.append(str(label))
            elif item:
                labels.append(str(item))
        return labels

    if isinstance(parsed_result, dict):
        return [str(key) for key, value in parsed_result.items() if value]

    if parsed_result:
        return [str(parsed_result)]
    return []


def serialize_detected_classes(raw_result):
    if raw_result is None:
        return '[]'
    if isinstance(raw_result, str):
        return raw_result
    return json.dumps(raw_result, ensure_ascii=False, indent=2)


def create_report_entry(
    file_path,
    nudity_detected,
    raw_result,
    confidence_score=0.0,
    media_type=None,
    model_name='',
    threshold_percent=60,
    thumbnail=None,
):
    """Create a report entry for a classified file.
    
    Args:
        file_path: Path to classified file
        nudity_detected: Whether nudity was detected
        raw_result: Raw detection result (classes/predictions)
        confidence_score: Confidence score (0-1)
        media_type: 'image', 'video', or None (auto-detected)
        model_name: Name of detection model used
        threshold_percent: Threshold used for detection
        thumbnail: Base64-encoded thumbnail or None
        
    Returns:
        Dictionary containing report entry data
    """
    normalized_confidence = max(0.0, min(float(confidence_score or 0.0), 1.0))
    return {
        'file': file_path,
        'media_type': media_type or detect_media_type(file_path),
        'model_name': model_name,
        'threshold_percent': threshold_to_percent(threshold_percent),
        'confidence_percent': round(normalized_confidence * 100, 2),
        'nudity_detected': bool(nudity_detected),
        'detected_classes': serialize_detected_classes(raw_result),
        'thumbnail': thumbnail or '',
        'date_classified': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
    }


def process_file(file_path, classify_image, classify_video):
    _, ext = os.path.splitext(file_path)
    ext = ext.lower()
    if ext in image_extensions:
        classify_image(file_path)
    elif ext in video_extensions:
        classify_video(file_path)
    else:
        logging.info('Skipping unsupported file: %s', file_path)


def process_file_queue(file_queue, classify_image, classify_video):
    """Worker thread: process files from queue.
    
    Args:
        file_queue: Queue of file paths
        classify_image: Image classification function
        classify_video: Video classification function
    """
    while not file_queue.empty():
        file_path = file_queue.get()
        try:
            process_file(file_path, classify_image, classify_video)
        except Exception as e:
            logging.error('Error processing file %s: %s', file_path, e)
        finally:
            file_queue.task_done()


def classify_files_in_folder(folder_path, classify_image, classify_video):
    """Classify all supported files in folder using worker threads.
    
    Args:
        folder_path: Root folder to scan
        classify_image: Image classification function
        classify_video: Video classification function
    """
    logging.debug('Starting classification in folder: %s', folder_path)
    file_queue = Queue()
    os.makedirs(DEFAULT_REPORT_DIR, exist_ok=True)

    for root, _, files in os.walk(folder_path):
        for file_name in files:
            file_queue.put(os.path.join(root, file_name))

    # Create non-daemon worker threads for proper resource cleanup
    workers = []
    for _ in range(10):
        worker = Thread(target=process_file_queue, args=(file_queue, classify_image, classify_video), daemon=False)
        worker.start()
        workers.append(worker)
    
    # Wait for all items to be processed
    file_queue.join()
    
    # Wait for all workers to complete with timeout
    for worker in workers:
        worker.join(timeout=5)
        if worker.is_alive():
            logging.warning('Worker thread did not complete within timeout')


def _write_report_rows(sheet, report_data):
    """Write report data rows to Excel sheet.
    
    Args:
        sheet: openpyxl worksheet
        report_data: List of report entry dictionaries
    """
    sheet.delete_rows(1, sheet.max_row or 1)
    sheet.append(REPORT_HEADERS)
    for data in report_data:
        sheet.append([
            data.get('file', ''),
            data.get('media_type', detect_media_type(data.get('file', ''))),
            data.get('model_name', ''),
            data.get('threshold_percent', 60),
            data.get('confidence_percent', 0),
            data.get('nudity_detected', False),
            data.get('detected_classes', '[]'),
            data.get('thumbnail', ''),
            data.get('date_classified', ''),
        ])


def save_scan_session(session_state, file_path):
    os.makedirs(os.path.dirname(file_path) or '.', exist_ok=True)
    with open(file_path, 'w', encoding='utf-8') as session_file:
        json.dump(session_state, session_file, ensure_ascii=False, indent=2)


def _embed_thumbnails_in_sheet(sheet, report_data):
    """Embed thumbnail images into report sheet thumbnail column.
    
    Args:
        sheet: openpyxl worksheet
        report_data: List of report entry dictionaries with thumbnail data
    """
    try:
        from openpyxl.drawing.image import Image as XLImage
    except ImportError:
        logging.debug('Could not import openpyxl Image; thumbnails will not be embedded in Excel')
        return
    
    if Image is None:
        logging.debug('PIL not available; thumbnails will not be embedded in Excel')
        return
    
    # Find Thumbnail column index
    headers = [cell.value for cell in sheet[1]]
    if 'Thumbnail' not in headers:
        logging.warning('Thumbnail column not found in report sheet')
        return
    
    thumbnail_col_idx = headers.index('Thumbnail') + 1  # Excel is 1-indexed
    thumbnail_col = openpyxl.utils.get_column_letter(thumbnail_col_idx)
    
    # Set column width and row height for thumbnails
    sheet.column_dimensions[thumbnail_col].width = 15
    
    # Embed thumbnails in rows
    for row_idx, entry in enumerate(report_data, start=2):  # Start at row 2 (after header)
        thumbnail_b64 = entry.get('thumbnail', '')
        if not thumbnail_b64:
            continue
        
        try:
            # Decode base64 to bytes
            thumbnail_data = base64.b64decode(thumbnail_b64)
            
            # Create PIL image and save to BytesIO
            img = Image.open(BytesIO(thumbnail_data))
            
            # Save to temporary BytesIO for openpyxl
            img_bytes = BytesIO()
            img.save(img_bytes, format='PNG')
            img_bytes.seek(0)
            
            # Insert image into cell
            cell_ref = f'{thumbnail_col}{row_idx}'
            xl_image = XLImage(img_bytes)
            xl_image.width = 100
            xl_image.height = 100
            sheet.add_image(xl_image, cell_ref)
            
            # Adjust row height to fit image
            sheet.row_dimensions[row_idx].height = 105
            
        except Exception as e:
            logging.warning('Failed to embed thumbnail in cell %s: %s', cell_ref, e)
            continue


def save_nudity_report(report_data, file_path, session_state=None):
    """Save report to Excel file with embedded thumbnails.
    
    Args:
        report_data: List of report entry dictionaries  
        file_path: Path to save Excel report
        session_state: Session metadata (auto-created if None)
    """
    os.makedirs(os.path.dirname(file_path) or '.', exist_ok=True)

    if os.path.exists(file_path):
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
    else:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Nudity Report'

    _write_report_rows(sheet, report_data)
    _embed_thumbnails_in_sheet(sheet, report_data)

    if 'Session' not in workbook.sheetnames:
        workbook.create_sheet('Session')
    session_sheet = workbook['Session']
    session_sheet.delete_rows(1, session_sheet.max_row or 1)

    if session_state is None:
        session_state = create_session_state(results=get_detected_results(report_data))

    session_sheet['A1'] = json.dumps(session_state, ensure_ascii=False, indent=2)
    workbook.save(file_path)

    save_scan_session(session_state, get_session_path(file_path))
    logging.info('Report saved to %s', file_path)


def check_and_save_report(report_data, file_path, batch_size=500, session_state=None):
    if len(report_data) >= batch_size:
        save_nudity_report(report_data, file_path, session_state=session_state)


def _sheet_headers(sheet):
    return [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]


def _legacy_row_to_entry(row):
    """Convert legacy report row to new report entry format.
    
    Args:
        row: Legacy report row tuple
        
    Returns:
        Report entry dictionary with new schema
    """
    detected_classes = row[2] if len(row) > 2 else '[]'
    return {
        'file': row[0],
        'media_type': detect_media_type(row[0] or ''),
        'model_name': '',
        'threshold_percent': 60,
        'confidence_percent': 0,
        'nudity_detected': bool(row[1]),
        'detected_classes': detected_classes or '[]',
        'thumbnail': '',
        'date_classified': row[3] if len(row) > 3 else '',
    }


def load_report_entries(file_path):
    if not os.path.exists(file_path):
        return []

    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    headers = _sheet_headers(sheet)
    normalized_headers = {str(header).strip(): index for index, header in enumerate(headers) if header}
    uses_new_schema = set(REPORT_HEADERS).issubset(set(normalized_headers))
    entries = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue

        if uses_new_schema:
            # Handle both old schema (Copied File) and new schema (Thumbnail)
            thumbnail_value = ''
            if 'Thumbnail' in normalized_headers:
                thumbnail_value = row[normalized_headers['Thumbnail']] or ''
            elif 'Copied File' in normalized_headers:
                # Migration from old schema: use empty thumbnail
                thumbnail_value = ''
            
            entry = {
                'file': row[normalized_headers['File']],
                'media_type': row[normalized_headers['Media Type']] or detect_media_type(row[normalized_headers['File']]),
                'model_name': row[normalized_headers['Model']] or '',
                'threshold_percent': row[normalized_headers['Threshold Percent']] or 60,
                'confidence_percent': row[normalized_headers['Confidence Percent']] or 0,
                'nudity_detected': bool(row[normalized_headers['Nudity Detected']]),
                'detected_classes': row[normalized_headers['Detected Classes']] or '[]',
                'thumbnail': thumbnail_value,
                'date_classified': row[normalized_headers['Date Classified']] or '',
            }
        else:
            entry = _legacy_row_to_entry(row)

        entries.append(entry)

    return entries


def load_existing_report(file_path):
    return {entry.get('file') for entry in load_report_entries(file_path)}


def load_scan_session(file_path):
    session_path = file_path if file_path.endswith('.json') else get_session_path(file_path)
    if os.path.exists(session_path):
        with open(session_path, 'r', encoding='utf-8') as session_file:
            return json.load(session_file)

    if os.path.exists(file_path):
        workbook = openpyxl.load_workbook(file_path)
        if 'Session' in workbook.sheetnames:
            raw_session = workbook['Session']['A1'].value
            if raw_session:
                try:
                    return json.loads(raw_session)
                except json.JSONDecodeError:
                    logging.warning('Could not parse embedded session state from %s', file_path)

        return create_session_state(results=get_detected_results(load_report_entries(file_path)))

    return create_session_state()


def handle_results(
    file_path,
    nudity_detected,
    raw_result,
    confidence_score=0.0,
    media_type=None,
    model_name='',
    threshold_percent=60,
    report_dir=DEFAULT_REPORT_DIR,
):
    """Handle detection results: create report entry and generate thumbnail.
    
    No longer copies files to report_dir. Files remain at original location.
    Thumbnails are generated for detected items (if dependencies available).
    
    Args:
        file_path: Original file path
        nudity_detected: Whether nudity was detected
        raw_result: Raw detection result
        confidence_score: Confidence score (0-1)
        media_type: Media type ('image'/'video'/None)
        model_name: Detection model name
        threshold_percent: Detection threshold
        report_dir: Directory for reports (not file copies)
        
    Returns:
        Report entry dictionary
    """
    # Generate thumbnail for detected items
    thumbnail = ''
    if nudity_detected:
        media_type = media_type or detect_media_type(file_path)
        thumbnail = get_thumbnail(file_path, media_type) or ''
    
    # Ensure report directory exists
    if nudity_detected:
        os.makedirs(report_dir, exist_ok=True)

    with report_lock:
        report_entry = create_report_entry(
            file_path,
            nudity_detected,
            raw_result,
            confidence_score=confidence_score,
            media_type=media_type,
            model_name=model_name,
            threshold_percent=threshold_percent,
            thumbnail=thumbnail,
        )
        nudity_report.append(report_entry)
        check_and_save_report(nudity_report, get_report_path(report_dir))
        return report_entry


def remove_report_entry(file_path):
    with report_lock:
        original_length = len(nudity_report)
        nudity_report[:] = [entry for entry in nudity_report if entry.get('file') != file_path]
        return len(nudity_report) != original_length


def delete_file_safely(file_path):
    if not os.path.exists(file_path):
        return False, 'File does not exist.'

    try:
        if send2trash is not None:
            send2trash(file_path)
            return True, 'Moved to trash.'

        if os.path.isdir(file_path):
            shutil.rmtree(file_path)
        else:
            os.remove(file_path)
        return True, 'Deleted permanently because trash support is unavailable.'
    except Exception as error:
        logging.error('Could not delete %s: %s', file_path, error)
        return False, str(error)


def open_file_location(file_path):
    target_path = file_path if os.path.isdir(file_path) else os.path.dirname(file_path)
    if not target_path:
        target_path = '.'

    try:
        if sys.platform == 'win32':
            os.startfile(target_path)
        elif sys.platform == 'darwin':
            subprocess.run(['open', target_path], check=False)
        else:
            subprocess.run(['xdg-open', target_path], check=False)
        return True, ''
    except Exception as error:
        logging.error('Could not open path %s: %s', target_path, error)
        return False, str(error)
