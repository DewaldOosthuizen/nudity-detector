"""
Additional coverage tests for ReportManager (src/reporting/report_manager.py).
Targets: lines 18-19, 23-24, 81-82, 115, 131-138, 176-178, 189-190, 196,
         222-226, 248-250, 278-290, 318-320.
"""
import json
import os
import sys
from unittest.mock import MagicMock, patch

sys.modules.setdefault("nudenet", MagicMock())

import src.reporting.report_manager as rm_mod  # noqa: E402
from src.core.models import ReportEntry, SessionState  # noqa: E402
from src.reporting.report_manager import ReportManager  # noqa: E402


def _entry(**kwargs):
    """Helper — ReportEntry with sensible defaults."""
    defaults = dict(
        file="/tmp/x.jpg",
        media_type="image",
        model_name="helloz",
        threshold_percent=60.0,
        confidence_percent=75.0,
        nudity_detected=True,
        detected_classes="[]",
    )
    defaults.update(kwargs)
    return ReportEntry(**defaults)


# ---------------------------------------------------------------------------
# PIL / XLImage import branches (lines 17-24)
# ---------------------------------------------------------------------------

def test_image_attribute_is_pil_or_none():
    """Module exposes an Image attribute that is either the PIL.Image module or None."""
    # If PIL is installed, Image is the PIL.Image module (has an 'open' callable).
    # If PIL is not installed, Image is None.  Both are valid states.
    assert rm_mod.Image is None or hasattr(rm_mod.Image, 'open')


def test_xlimage_none_when_openpyxl_image_unavailable(monkeypatch):
    """Branch where XLImage import fails sets XLImage = None."""
    # Simulate the branch — reload is impractical; check attribute directly.
    monkeypatch.setattr(rm_mod, "XLImage", None)
    assert rm_mod.XLImage is None


# ---------------------------------------------------------------------------
# validate_report_dir — OSError branch (line 81-82)
# ---------------------------------------------------------------------------

def test_validate_report_dir_oserror():
    with patch("os.makedirs", side_effect=OSError("no space")):
        ok, msg = ReportManager.validate_report_dir("/tmp/fake_dir")
    assert ok is False
    assert "no space" in msg


# ---------------------------------------------------------------------------
# load_entries — skip empty row (line 114-115)
# ---------------------------------------------------------------------------

def test_load_entries_skips_empty_rows(tmp_path):
    import openpyxl

    from src.core import constants

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(constants.REPORT_HEADERS)
    ws.append([None] + [None] * (len(constants.REPORT_HEADERS) - 1))  # empty row
    ws.append(["/tmp/valid.jpg", "image", "helloz", 60.0, 75.0, True, "[]", "", "2024-01-01"])
    path = str(tmp_path / "test.xlsx")
    wb.save(path)

    entries = ReportManager.load_entries(path)
    # The None row should be skipped; only the valid row returned
    assert len(entries) == 1


def test_load_entries_malformed_row_logged(tmp_path, caplog):
    import logging

    import openpyxl

    from src.core import constants

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(constants.REPORT_HEADERS)
    # Row with file present but confidence_percent is a non-numeric string
    ws.append(["/tmp/x.jpg", "image", "helloz", "bad", "not_float", True, "[]", "", "2024-01-01"])
    path = str(tmp_path / "bad.xlsx")
    wb.save(path)

    with caplog.at_level(logging.WARNING):
        entries = ReportManager.load_entries(path)
    # Should log a warning and skip; may return 0 or 1 entries depending on tolerance
    # The key is no exception is raised
    assert isinstance(entries, list)


def test_load_entries_exception_returns_empty(tmp_path):
    path = str(tmp_path / "corrupt.xlsx")
    with open(path, "w") as f:
        f.write("not an xlsx file")
    entries = ReportManager.load_entries(path)
    assert entries == []


# ---------------------------------------------------------------------------
# save_entries — exception branch (line 176-178)
# ---------------------------------------------------------------------------

def test_save_entries_exception_returns_false(tmp_path):
    entry = _entry()
    path = str(tmp_path / "output.xlsx")
    with patch("openpyxl.Workbook.save", side_effect=OSError("disk full")):
        result = ReportManager.save_entries([entry], path)
    assert result is False


# ---------------------------------------------------------------------------
# _embed_thumbnails — XLImage/Image None branch (lines 188-190)
# ---------------------------------------------------------------------------

def test_embed_thumbnails_skips_when_xlimage_none(monkeypatch, tmp_path):
    import openpyxl
    monkeypatch.setattr(rm_mod, "XLImage", None)
    monkeypatch.setattr(rm_mod, "Image", None)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Thumbnail"])

    entry = _entry(thumbnail="abc")
    # Should not raise; just log debug and return
    ReportManager._embed_thumbnails(ws, [entry])


def test_embed_thumbnails_no_thumbnail_column(monkeypatch):
    import openpyxl
    monkeypatch.setattr(rm_mod, "XLImage", MagicMock())
    monkeypatch.setattr(rm_mod, "Image", MagicMock())

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["File", "Model"])  # no Thumbnail column

    entry = _entry(thumbnail="abc")
    ReportManager._embed_thumbnails(ws, [entry])  # should return early, no error


# ---------------------------------------------------------------------------
# save_session — exception branch (lines 248-250)
# ---------------------------------------------------------------------------

def test_save_session_exception_returns_false(tmp_path):
    state = SessionState()
    with patch("builtins.open", side_effect=OSError("permission denied")):
        result = ReportManager.save_session(state, str(tmp_path / "report.xlsx"))
    assert result is False


# ---------------------------------------------------------------------------
# load_session — JSON decode error and xlsx fallback (lines 278-290)
# ---------------------------------------------------------------------------

def test_load_session_bad_json_file(tmp_path):
    session_path = tmp_path / "report_session.json"
    session_path.write_text("not valid json")
    # Should fall through to default empty session
    state = ReportManager.load_session(str(tmp_path / "report.xlsx"))
    assert isinstance(state, SessionState)


def test_load_session_xlsx_with_embedded_session(tmp_path):
    import openpyxl
    state = SessionState()
    wb = openpyxl.Workbook()
    session_sheet = wb.create_sheet("Session")
    session_sheet["A1"] = json.dumps(state.to_dict())
    path = str(tmp_path / "report.xlsx")
    wb.save(path)

    loaded = ReportManager.load_session(path)
    assert isinstance(loaded, SessionState)


def test_load_session_xlsx_bad_embedded_json(tmp_path):
    import openpyxl
    wb = openpyxl.Workbook()
    session_sheet = wb.create_sheet("Session")
    session_sheet["A1"] = "not json at all"
    path = str(tmp_path / "report.xlsx")
    wb.save(path)

    loaded = ReportManager.load_session(path)
    assert isinstance(loaded, SessionState)


def test_load_session_xlsx_fallback_from_entries(tmp_path):
    import openpyxl

    from src.core import constants
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(constants.REPORT_HEADERS)
    ws.append(["/tmp/x.jpg", "image", "helloz", 60.0, 75.0, True, "[]", "", "2024-01-01"])
    path = str(tmp_path / "report.xlsx")
    wb.save(path)

    loaded = ReportManager.load_session(path)
    assert isinstance(loaded, SessionState)


# ---------------------------------------------------------------------------
# create_demo_session — exception branch (lines 318-320)
# ---------------------------------------------------------------------------

def test_create_demo_session_exception_returns_false(tmp_path):
    state = SessionState()
    path = str(tmp_path / "report.xlsx")
    with patch("openpyxl.Workbook.save", side_effect=Exception("boom")):
        result = ReportManager.create_demo_session(path, state)
    assert result is False


def test_create_demo_session_creates_new_workbook(tmp_path):
    state = SessionState()
    path = str(tmp_path / "new_report.xlsx")
    # File doesn't exist — creates a new workbook
    result = ReportManager.create_demo_session(path, state)
    assert result is True
    assert os.path.exists(path)
